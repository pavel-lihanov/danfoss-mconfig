import openpyxl
import pickle
import zipfile
from mconfig.models import Order, Profile,Option_prices,Base_price
import devices
from tkinter import *
from django.contrib.messages import constants as messages
from django.contrib import messages
from django.template import loader
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseForbidden, HttpResponseRedirect, HttpResponseServerError
from django.http import HttpResponseRedirect
from django.shortcuts import render_to_response

SALE_PRICE_COLUMN = 21
SUPPLIER_PRICE_COLUMN = 20
NAME_COLUMN = 1
POWER_COLUMN = 2
VOLTAGE_COLUMN = 4 

class PriceNotSpecified(Exception):
    pass

class NotInPricelist(Exception):
    pass


    
        
class PricingError(Exception):
    pass
    
class Price:
    #price struct contains price details
    supplier_price = 0.0
    sale_price = 0.0
    delivery_cost = 0.0
    def __init__(self):
        pass
        
    @property
    def total(self):
        total = self.sale_price + self.delivery_cost
        return total
    
class PriceList:    
    def get_price(self, package):
        raise PricingError('Abstract method')
        
class VEDACachedPriceList(PriceList):
    def __init__(self, cached_path):
        with zipfile.open(cached_path) as zip:
            with zip.open('price.bin') as f:
                self.prices = pickle.loads(f.read())
            
    @staticmethod
    def create_cache(pricelist_path, cached_path):
        cache = {}
        wb = openpyxl.load_workbook(pricelist_path)
        pricelist = wb.active
        for row in pricelist.iter_rows(min_row=1):
            try:
                #make code from row
                try:                    
                    #code = '{0}-{1}{2}{3}{4}{5}{6}{7}{8:3.0f}{9}{10}A{11}B{12}C{13}D{14}{15}E{16}{17}'.format(
                    code = '{0}-{1}{2}{3}{4}{5}{6}{7}{8:03.0f}{9}{10}A{11}B{12}C{13}D{14}E{15}{16}'.format(
                                           row[NAME_COLUMN].value, 
                                           row[POWER_COLUMN].value,     #power
                                           row[VOLTAGE_COLUMN].value,   #voltage
                                           row[5].value,                #frequency
                                           row[6].value,                #protection
                                           row[7].value,                #motor type
                                           row[8].value,                #control mode                                       
                                           row[9].value,                #braking
                                           row[10].value,               #power cell current
                                           row[11].value,               #cooling
                                           row[12].value,               #power cell bypass
                                           row[13].value,               #A
                                           row[14].value,               #B
                                           row[15].value,               #C
                                           row[16].value,               #D
                                           row[17].value,               #mains location
                                           row[18].value,               #E
                                           row[19].value,               #service access
                                           )                    
                except TypeError:
                    raise
                supp_price = float(row[SUPPLIER_PRICE_COLUMN].value)
                sale_price = float(row[SALE_PRICE_COLUMN].value)            
                cache[code] = (supp_price, sale_price)
            except (TypeError, ValueError) as ex:
                #print('Failed to convert row ', code, row[SUPPLIER_PRICE_COLUMN].value, row[SALE_PRICE_COLUMN].value)
                pass
        
        with zipfile.ZipFile(cached_path, 'w', compression=zipfile.ZIP_BZIP2) as zip:
            zip.writestr('price.bin', pickle.dumps(cache))
        
    def get_price(self, package):
        order_code = package.order_code
        if order_code in self.prices:
            p = Price()
            p.supplier_price, p.sale_price = self.prices[order_code]

            return p

        else:
            raise NotInPricelist('{0}: not found in pricelist'.format(package.name))
   
class VEDAXLPriceList(PriceList):
    def __init__(self, pricelist_path):
        self.wb = openpyxl.load_workbook(pricelist_path)
        #TODO: store version in xls file
        self.version = '0.0'
        
        
    def get_price(self, package):
        pricelist = self.wb.active
        for r in pricelist.iter_rows(min_row=1):
            if self.package_matches_pricelist(r, package):

                
                
                #базовая цена без учета опций и без учета доставки
                base_set=Base_price.objects.filter(nom_voltage=package.attributes['voltage']/1000,current=package.attributes['nom_current']).values('price')
                print('PRICE=')
                for i in base_set:
                    price_test=i['price']
                    print (price_test)
                    
                def execute_opt(self):
                    opt=Option_prices.objects.filter(option_value=package.options[self]).values('calc_data')
                    for i in opt:
                        result=i['calc_data']
                        #print(result)
                    return result
                    
                #calc_data для корпуса
                enclosure_calc=execute_opt('enclosure')
                print('enclosure_calc=')
                print(enclosure_calc)
                #calc_data для корпуса
                #opt_corp=Option_prices.objects.filter(option_value=package.options['enclosure']).values('calc_data')
                #print('opt_corp,calc_data=')
                #for i in opt_corp:
                #    enclosure_calc_data=i['calc_data']   
                #    print (enclosure_calc_data)
     
                byp_opt=execute_opt('power_cell_autobypass')
                print('byp_opt=')
                print(byp_opt)
                
                cell=execute_opt('power_cells')
                print('power_cells=')
                print(cell)
                 
                fieldbus=execute_opt('fieldbus')
                print('fieldbus=')
                print(fieldbus)
                
                control_opt=execute_opt('control_mode')
                print('control_opt=')
                print(control_opt) 
                
                cell_break=execute_opt('brake_mode')
                print('cell_break=')
                print(cell_break)
                
                service=execute_opt('Service access')
                print('Service access=')
                print(service)
                 
                #итоговая цена с учетом опций
                price_test_with_option=price_test+price_test*enclosure_calc+price_test*byp_opt+price_test*cell+ fieldbus+price_test*control_opt+price_test*cell_break+price_test*service
                print('price_test_with_option=')
                print(price_test_with_option)
                price_test=float(price_test_with_option)
                
                price_cell = r[SALE_PRICE_COLUMN]
                supp_price_cell = r[SUPPLIER_PRICE_COLUMN]
                #price_cell =r[price_test]
                #supp_price_cell=r[price_test]
                if type(price_cell.value) is float:
                    p = Price()
                    p.sale_price = price_cell.value
                    p.supplier_price = supp_price_cell.value
                    return p
                    print('PRICE=')
                    print(p)
                    return '{0:.2f}'.format(price_cell.value)                    
                else:
                   raise PriceNotSpecified('Unexpected price value:', price_cell.value)
                #if type(price_cell.value) is float:
                #    p = Price()
                #    p.sale_price = price_cell.value
                #    p.supplier_price = supp_price_cell.value
                 #   return p
                #    return '{0:.2f}'.format(price_cell.value)
                #else:
                #   raise PriceNotSpecified('Unexpected price value:', price_cell.value)    
        raise NotInPricelist('{0}: not found in pricelist'.format(package.name))

        
    def package_matches_pricelist(self, row, package):                
        price_name = '{0}-{1}{2}'.format(row[NAME_COLUMN].value, row[POWER_COLUMN].value, row[VOLTAGE_COLUMN].value)
        #price_name = '{0}-{1}{2}'.format(row['VDD'], row[POWER_COLUMN].value, row[VOLTAGE_COLUMN].value)
        if package.name != price_name:
            return False
            
        for name,val in package.options.infos():
            if not val.matches_pricelist(package.options[name], row):
                return False
        
        return True
price_lists = {}               


class VEDADBPriceList(PriceList):

    def __init__(self):
        pass    
    def get_price(self, package):
        try:
            #базовая цена без учета опций и без учета доставки
            base_set=Base_price.objects.filter(nom_voltage=package.attributes['voltage']/1000,current=package.attributes['nom_current']).values('price')
            print('PRICE=')
            for i in base_set:
                price_test=i['price']
                print (price_test)
                
            def execute_opt(self):
                opt=Option_prices.objects.filter(option_value=package.options[self]).values('calc_data')
                for i in opt:
                    result=i['calc_data']
                    #print(result)
                return result
                
            #calc_data для корпуса
            enclosure_calc=execute_opt('enclosure')
            print('enclosure_calc=')
            print(enclosure_calc)

            byp_opt=execute_opt('power_cell_autobypass')
            print('byp_opt=')
            print(byp_opt)
            
            cell=execute_opt('power_cells')
            print('power_cells=')
            print(cell)
             
            fieldbus=execute_opt('fieldbus')
            print('fieldbus=')
            print(fieldbus)
            
            control_opt=execute_opt('control_mode')
            print('control_opt=')
            print(control_opt) 
            
            cell_break=execute_opt('brake_mode')
            print('cell_break=')
            print(cell_break)
            
            service=execute_opt('Service access')
            print('Service access=')
            print(service)
             
            #итоговая цена с учетом опций
            price_test_with_option=price_test+price_test*enclosure_calc+price_test*byp_opt+price_test*cell+ fieldbus+price_test*control_opt+price_test*cell_break+price_test*service
            print('price_test_with_option=')
            print(price_test_with_option)

            price_cell =float(price_test_with_option)
            print ('price_cell=')
            print (price_cell)
            supp_price_cell=float(price_test_with_option)
            print ('supp_price_cell=')
            print (supp_price_cell)
            p = Price()
            p.sale_price = price_cell
            p.supplier_price = supp_price_cell
            return p
        except Exception as exc:
            def my_info(request):
                messages.info(request, 'Цены нет!!!!')
                return render_to_response('result.html')
            
            
            raise NotInPricelist('{0}: not found in pricelist'.format(package.name))
            

            
